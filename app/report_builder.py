#!/usr/bin/env python3
"""
build_test_scenarios_output.py

Turns requirement-validation + test-scenario-generation results into:
  1. A styled HTML report (validation summary, test scenarios, traceability matrix)
  2. An Excel workbook (.xlsx) with 3 sheets: Validation, Test Scenarios, Traceability Matrix

This is the output stage of the "Req2QA" workflow:
  application name + requirements file (docx/xlsx/pdf)
    -> Claude validates requirements against the named application
    -> Claude generates test scenarios per valid requirement
    -> this script renders both deliverables

No API key is required anywhere in this workflow - the validation and
generation reasoning is done by Claude directly in the chat session, not
via a separate API call. This script only formats the results Claude
already produced.

USAGE
-----
    python3 build_test_scenarios_output.py data.json report_out.html data_out.xlsx

INPUT data.json SHAPE
----------------------
{
  "application": "Salesforce Sales Cloud",
  "requirements_source": "requirements.docx",
  "run_date": "2026-09-02 15:10",
  "validation": [
    {
      "req_id": "REQ-001",
      "requirement": "System shall allow login via SSO",
      "valid_for_app": true,
      "notes": ""
    },
    {
      "req_id": "REQ-007",
      "requirement": "System shall support offline mode for 30 days",
      "valid_for_app": false,
      "notes": "Salesforce Sales Cloud has no native 30-day offline mode; needs clarification or a mobile-offline add-on."
    }
  ],
  "test_scenarios": [
    {
      "tc_id": "TC-001",
      "req_id": "REQ-001",
      "title": "Verify SSO login succeeds with valid identity provider session",
      "precondition": "User has an active SSO session with the configured IdP",
      "steps": "1. Navigate to login URL\\n2. Click 'Login with SSO'\\n3. Complete IdP redirect",
      "expected_result": "User is authenticated and lands on their home page"
    }
  ]
}

Requirements with valid_for_app = false are excluded from test generation
by design (flagged, not scenario'd) but still appear in the Validation
sheet/section and in the traceability matrix as NOT_TESTABLE.
"""
import json
import sys
import html
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required: pip install openpyxl --break-system-packages")
    raise


def esc(s):
    return html.escape(str(s if s is not None else ""))


def nl2br(s):
    return esc(s).replace("\\n", "<br>").replace("\n", "<br>")


# --------------------------------------------------------------------------- HTML
def build_html(data):
    validation = data.get("validation", [])
    scenarios = data.get("test_scenarios", [])

    total_reqs = len(validation)
    valid_reqs = sum(1 for v in validation if v.get("valid_for_app"))
    flagged_reqs = total_reqs - valid_reqs
    total_tcs = len(scenarios)

    # traceability: req_id -> list of tc_ids
    trace = defaultdict(list)
    for s in scenarios:
        trace[s.get("req_id", "")].append(s.get("tc_id", ""))

    val_rows = []
    for v in validation:
        ok = v.get("valid_for_app")
        badge = '<span class="badge valid">VALID</span>' if ok else '<span class="badge flagged">FLAGGED</span>'
        row_cls = "row-valid" if ok else "row-flagged"
        val_rows.append(
            f'<tr class="{row_cls}"><td><strong>{esc(v.get("req_id",""))}</strong></td>'
            f'<td>{esc(v.get("requirement",""))}</td><td>{badge}</td>'
            f'<td>{esc(v.get("notes",""))}</td></tr>'
        )
    val_html = "\n".join(val_rows) if val_rows else '<tr><td colspan="4" class="empty">No requirements found.</td></tr>'

    tc_rows = []
    for s in scenarios:
        tc_rows.append(
            f'<tr><td><strong>{esc(s.get("tc_id",""))}</strong></td>'
            f'<td>{esc(s.get("req_id",""))}</td>'
            f'<td>{esc(s.get("title",""))}</td>'
            f'<td>{esc(s.get("precondition",""))}</td>'
            f'<td>{nl2br(s.get("steps",""))}</td>'
            f'<td>{esc(s.get("expected_result",""))}</td></tr>'
        )
    tc_html = "\n".join(tc_rows) if tc_rows else '<tr><td colspan="6" class="empty">No test scenarios generated.</td></tr>'

    trace_rows = []
    for v in validation:
        rid = v.get("req_id", "")
        tcs = trace.get(rid, [])
        if not v.get("valid_for_app"):
            status = '<span class="badge flagged">NOT TESTABLE</span>'
        elif tcs:
            status = '<span class="badge valid">COVERED</span>'
        else:
            status = '<span class="badge gap">GAP</span>'
        trace_rows.append(
            f'<tr><td><strong>{esc(rid)}</strong></td><td>{esc(v.get("requirement",""))}</td>'
            f'<td class="tc-ids">{esc(", ".join(tcs) if tcs else "-")}</td><td>{status}</td></tr>'
        )
    trace_html = "\n".join(trace_rows) if trace_rows else '<tr><td colspan="4" class="empty">No traceability data.</td></tr>'

    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Test Scenarios &amp; Traceability Report</title><style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{--navy:#0F2044;--blue:#1A5EA8;--sky:#E8F1FB;--green:#15803D;--gbg:#DCFCE7;
--red:#B91C1C;--rbg:#FEE2E2;--amber:#B45309;--abg:#FEF3C7;--purple:#6D28D9;
--pbg:#EDE9FE;--border:#E2E8F0;--text:#1E293B;--muted:#64748B;--bg:#F8FAFC}}
body{{font-family:system-ui,sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.6}}
.wrap{{max-width:1300px;margin:0 auto;padding:32px 24px 64px}}
.hdr{{background:var(--navy);color:#fff;border-radius:12px;padding:36px 40px;margin-bottom:28px}}
.hdr h1{{font-size:24px;font-weight:700}}
.hdr .sub{{color:#94A3B8;font-size:13px;margin-top:6px}}
.hdr .meta{{display:flex;gap:32px;margin-top:20px;flex-wrap:wrap}}
.mi{{font-size:12px;color:#CBD5E1}}.mi strong{{display:block;color:#fff;font-size:13px;margin-bottom:2px}}
.sc{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:32px}}
.card{{background:#fff;border:1px solid var(--border);border-radius:10px;padding:20px 14px;text-align:center}}
.card .n{{font-size:34px;font-weight:700;line-height:1.1}}.card .l{{font-size:12px;color:var(--muted);margin-top:4px}}
.card.good .n{{color:var(--green)}}.card.bad .n{{color:var(--red)}}
.card.warn .n{{color:var(--amber)}}.card.info .n{{color:var(--blue)}}
.st{{font-size:16px;font-weight:700;color:var(--navy);margin:36px 0 14px;padding-bottom:10px;
border-bottom:2px solid var(--border);display:flex;align-items:center;gap:10px}}
.pill{{font-size:11px;font-weight:600;background:var(--sky);color:var(--blue);padding:2px 10px;border-radius:20px}}
.tw{{overflow-x:auto;border-radius:10px;border:1px solid var(--border);margin-bottom:8px}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:13px}}
thead th{{background:var(--navy);color:#fff;padding:11px 14px;text-align:left;font-weight:600;font-size:12px;white-space:nowrap}}
thead th:first-child{{border-radius:9px 0 0 0}}thead th:last-child{{border-radius:0 9px 0 0}}
tbody tr{{border-bottom:1px solid var(--border)}}tbody tr:last-child{{border-bottom:none}}
tbody tr:hover{{background:#F1F5F9}}tbody td{{padding:10px 14px;vertical-align:top}}
.row-valid{{background:#FAFFF9}}.row-flagged{{background:#FFFDF0}}
.tc-ids{{font-family:monospace;font-size:12px;color:var(--blue)}}
.empty{{text-align:center;color:var(--muted);padding:24px;font-style:italic}}
.badge{{display:inline-block;font-size:11px;font-weight:600;padding:2px 9px;border-radius:4px;white-space:nowrap}}
.badge.valid{{background:var(--gbg);color:var(--green)}}.badge.flagged{{background:var(--abg);color:var(--amber)}}
.badge.gap{{background:var(--rbg);color:var(--red)}}
.ft{{margin-top:48px;text-align:center;font-size:12px;color:var(--muted);
border-top:1px solid var(--border);padding-top:20px}}
@media print{{.tw{{overflow:visible}}body{{background:#fff}}}}
</style></head><body><div class="wrap">
<header class="hdr"><h1>Test Scenarios &amp; Requirements Traceability Report</h1>
<p class="sub">Application-aware requirement validation and scenario generation</p>
<div class="meta">
<div class="mi"><strong>Application</strong>{esc(data.get("application",""))}</div>
<div class="mi"><strong>Requirements Source</strong>{esc(data.get("requirements_source",""))}</div>
<div class="mi"><strong>Run Date</strong>{esc(data.get("run_date",""))}</div>
{f'<div class="mi"><strong>Baseline</strong>{esc(data.get("baseline_version"))}</div>' if data.get("baseline_version") else ""}
</div></header>

<div class="sc">
<div class="card info"><div class="n">{total_reqs}</div><div class="l">Total Requirements</div></div>
<div class="card good"><div class="n">{valid_reqs}</div><div class="l">Valid for Application</div></div>
<div class="card warn"><div class="n">{flagged_reqs}</div><div class="l">Flagged / Needs Review</div></div>
<div class="card good"><div class="n">{total_tcs}</div><div class="l">Test Scenarios Generated</div></div>
</div>

<h2 class="st">Requirement Validation <span class="pill">TABLE 1</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:100px">Req ID</th><th style="width:40%">Requirement</th>
<th style="width:110px">Status</th><th>Notes</th>
</tr></thead><tbody>
{val_html}
</tbody></table></div>

<h2 class="st">Generated Test Scenarios <span class="pill">TABLE 2</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:90px">TC ID</th><th style="width:90px">Req ID</th>
<th style="width:22%">Title</th><th style="width:18%">Precondition</th>
<th>Steps</th><th style="width:18%">Expected Result</th>
</tr></thead><tbody>
{tc_html}
</tbody></table></div>

<h2 class="st">Requirement &harr; Test Case Traceability Matrix <span class="pill">TABLE 3</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:100px">Req ID</th><th style="width:35%">Requirement</th>
<th style="width:20%">Linked TC IDs</th><th style="width:130px">Status</th>
</tr></thead><tbody>
{trace_html}
</tbody></table></div>

<div class="ft">Req2QA &mdash; Requirements to Test Coverage &bull; {esc(data.get("run_date",""))}</div>
</div></body></html>"""


# --------------------------------------------------------------------------- XLSX
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GOOD_FILL = PatternFill("solid", fgColor="DCFCE7")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
BAD_FILL = PatternFill("solid", fgColor="FEE2E2")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_xlsx(data, out_path):
    wb = Workbook()

    # Sheet 1: Validation
    ws1 = wb.active
    ws1.title = "Validation"
    ws1.append(["Req ID", "Requirement", "Valid for Application", "Notes"])
    for v in data.get("validation", []):
        ok = v.get("valid_for_app")
        ws1.append([v.get("req_id", ""), v.get("requirement", ""), "YES" if ok else "FLAGGED", v.get("notes", "")])
        r = ws1.max_row
        fill = GOOD_FILL if ok else WARN_FILL
        for c in range(1, 5):
            ws1.cell(row=r, column=c).fill = fill
            ws1.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws1, 4)
    autosize(ws1, [12, 55, 20, 45])

    # Sheet 2: Test Scenarios
    ws2 = wb.create_sheet("Test Scenarios")
    ws2.append(["TC ID", "Req ID", "Title", "Precondition", "Steps", "Expected Result"])
    for s in data.get("test_scenarios", []):
        ws2.append([
            s.get("tc_id", ""), s.get("req_id", ""), s.get("title", ""),
            s.get("precondition", ""), s.get("steps", "").replace("\\n", "\n"),
            s.get("expected_result", ""),
        ])
        r = ws2.max_row
        for c in range(1, 7):
            ws2.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws2, 6)
    autosize(ws2, [10, 10, 32, 28, 45, 32])

    # Sheet 3: Traceability Matrix
    ws3 = wb.create_sheet("Traceability Matrix")
    ws3.append(["Req ID", "Requirement", "Linked TC IDs", "Status"])
    trace = defaultdict(list)
    for s in data.get("test_scenarios", []):
        trace[s.get("req_id", "")].append(s.get("tc_id", ""))
    for v in data.get("validation", []):
        rid = v.get("req_id", "")
        tcs = trace.get(rid, [])
        if not v.get("valid_for_app"):
            status, fill = "NOT_TESTABLE", WARN_FILL
        elif tcs:
            status, fill = "COVERED", GOOD_FILL
        else:
            status, fill = "GAP", BAD_FILL
        ws3.append([rid, v.get("requirement", ""), ", ".join(tcs), status])
        r = ws3.max_row
        for c in range(1, 5):
            ws3.cell(row=r, column=c).fill = fill
            ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws3, 4)
    autosize(ws3, [12, 55, 25, 16])

    wb.save(out_path)


# --------------------------------------------------------------------------- Option B (custom apps): 3-way traceability
def build_html_custom(data, flow):
    """Custom Application Mode report: consistency/testability validation
    plus three-way traceability (requirement <-> process step <-> test
    case), using the same generation output run_qa_analysis_custom()
    produces. Kept as a separate function from build_html() rather than
    branching inside it, per the design decision to keep Option A untouched."""
    validation = data.get("validation", [])
    scenarios = data.get("test_scenarios", [])
    flow_steps = flow.get("steps", [])
    uncovered_flow_steps = set(data.get("uncovered_flow_steps", []))

    total_reqs = len(validation)
    testable_reqs = sum(1 for v in validation if v.get("testable"))
    flagged_reqs = total_reqs - testable_reqs
    total_tcs = len(scenarios)
    total_flow_steps = len(flow_steps)

    trace = defaultdict(list)  # req_id -> [tc_id]
    for s in scenarios:
        trace[s.get("req_id", "")].append(s.get("tc_id", ""))

    val_rows = []
    for v in validation:
        ok = v.get("testable")
        badge = '<span class="badge valid">TESTABLE</span>' if ok else '<span class="badge flagged">FLAGGED</span>'
        row_cls = "row-valid" if ok else "row-flagged"
        step_ids = ", ".join(v.get("flow_step_ids", [])) or "-"
        val_rows.append(
            f'<tr class="{row_cls}"><td><strong>{esc(v.get("req_id",""))}</strong></td>'
            f'<td>{esc(v.get("requirement",""))}</td><td>{badge}</td>'
            f'<td class="tc-ids">{esc(step_ids)}</td><td>{esc(v.get("notes",""))}</td></tr>'
        )
    val_html = "\n".join(val_rows) if val_rows else '<tr><td colspan="5" class="empty">No requirements found.</td></tr>'

    tc_rows = []
    for s in scenarios:
        tc_rows.append(
            f'<tr><td><strong>{esc(s.get("tc_id",""))}</strong></td>'
            f'<td>{esc(s.get("req_id",""))}</td>'
            f'<td class="tc-ids">{esc(", ".join(s.get("flow_step_ids", [])) or "-")}</td>'
            f'<td>{esc(s.get("title",""))}</td>'
            f'<td>{esc(s.get("precondition",""))}</td>'
            f'<td>{nl2br(s.get("steps",""))}</td><td>{esc(s.get("expected_result",""))}</td></tr>'
        )
    tc_html = "\n".join(tc_rows) if tc_rows else '<tr><td colspan="7" class="empty">No test scenarios generated.</td></tr>'

    # Three-way trace: one row per requirement plus one row per otherwise-uncovered flow step
    trace_rows = []
    for v in validation:
        rid = v.get("req_id", "")
        tcs = trace.get(rid, [])
        step_ids = v.get("flow_step_ids", [])
        if not v.get("testable"):
            status = '<span class="badge flagged">NOT TESTABLE</span>'
        elif not step_ids:
            status = '<span class="badge gap">NO FLOW STEP</span>'
        elif tcs:
            status = '<span class="badge valid">COVERED</span>'
        else:
            status = '<span class="badge gap">GAP</span>'
        trace_rows.append(
            f'<tr><td><strong>{esc(rid)}</strong></td><td>{esc(v.get("requirement",""))}</td>'
            f'<td class="tc-ids">{esc(", ".join(step_ids) or "-")}</td>'
            f'<td class="tc-ids">{esc(", ".join(tcs) if tcs else "-")}</td><td>{status}</td></tr>'
        )
    step_lookup = {st.get("step_id", ""): st for st in flow_steps}
    for step_id in sorted(uncovered_flow_steps):
        st = step_lookup.get(step_id, {})
        trace_rows.append(
            f'<tr class="row-flagged"><td>-</td>'
            f'<td><em>{esc(st.get("screen_or_stage", step_id))}</em> - flow step with no requirement</td>'
            f'<td class="tc-ids">{esc(step_id)}</td><td class="tc-ids">-</td>'
            f'<td><span class="badge gap">UNREQUESTED STEP</span></td></tr>'
        )
    trace_html = "\n".join(trace_rows) if trace_rows else '<tr><td colspan="5" class="empty">No traceability data.</td></tr>'

    flow_rows = []
    for st in flow_steps:
        gap_badge = '<span class="badge gap">NO REQUIREMENT</span>' if st.get("step_id") in uncovered_flow_steps else '<span class="badge valid">COVERED</span>'
        flow_rows.append(
            f'<tr><td><strong>{esc(st.get("step_id",""))}</strong></td>'
            f'<td>{esc(st.get("screen_or_stage",""))}</td>'
            f'<td>{esc(st.get("description",""))}</td>'
            f'<td>{"Yes - " + esc(st.get("decision_detail","")) if st.get("decision_point") else "No"}</td>'
            f'<td>{gap_badge}</td></tr>'
        )
    flow_html = "\n".join(flow_rows) if flow_rows else '<tr><td colspan="5" class="empty">No process flow steps.</td></tr>'

    baseline_html = f'<div class="mi"><strong>Baseline</strong>{esc(data.get("baseline_version"))}</div>' if data.get("baseline_version") else ""

    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Custom Application - Test Coverage Report</title><style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{--navy:#0F2044;--blue:#1A5EA8;--sky:#E8F1FB;--green:#15803D;--gbg:#DCFCE7;
--red:#B91C1C;--rbg:#FEE2E2;--amber:#B45309;--abg:#FEF3C7;--purple:#6D28D9;
--pbg:#EDE9FE;--border:#E2E8F0;--text:#1E293B;--muted:#64748B;--bg:#F8FAFC}}
body{{font-family:system-ui,sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.6}}
.wrap{{max-width:1300px;margin:0 auto;padding:32px 24px 64px}}
.hdr{{background:var(--navy);color:#fff;border-radius:12px;padding:36px 40px;margin-bottom:28px}}
.hdr h1{{font-size:24px;font-weight:700}}
.hdr .sub{{color:#94A3B8;font-size:13px;margin-top:6px}}
.hdr .meta{{display:flex;gap:32px;margin-top:20px;flex-wrap:wrap}}
.mi{{font-size:12px;color:#CBD5E1}}.mi strong{{display:block;color:#fff;font-size:13px;margin-bottom:2px}}
.sc{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:32px}}
.card{{background:#fff;border:1px solid var(--border);border-radius:10px;padding:20px 14px;text-align:center}}
.card .n{{font-size:34px;font-weight:700;line-height:1.1}}.card .l{{font-size:12px;color:var(--muted);margin-top:4px}}
.card.good .n{{color:var(--green)}}.card.bad .n{{color:var(--red)}}
.card.warn .n{{color:var(--amber)}}.card.info .n{{color:var(--blue)}}
.st{{font-size:16px;font-weight:700;color:var(--navy);margin:36px 0 14px;padding-bottom:10px;
border-bottom:2px solid var(--border);display:flex;align-items:center;gap:10px}}
.pill{{font-size:11px;font-weight:600;background:var(--sky);color:var(--blue);padding:2px 10px;border-radius:20px}}
.tw{{overflow-x:auto;border-radius:10px;border:1px solid var(--border);margin-bottom:8px}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:13px}}
thead th{{background:var(--navy);color:#fff;padding:11px 14px;text-align:left;font-weight:600;font-size:12px;white-space:nowrap}}
thead th:first-child{{border-radius:9px 0 0 0}}thead th:last-child{{border-radius:0 9px 0 0}}
tbody tr{{border-bottom:1px solid var(--border)}}tbody tr:last-child{{border-bottom:none}}
tbody tr:hover{{background:#F1F5F9}}tbody td{{padding:10px 14px;vertical-align:top}}
.row-valid{{background:#FAFFF9}}.row-flagged{{background:#FFFDF0}}
.tc-ids{{font-family:monospace;font-size:12px;color:var(--blue)}}
.empty{{text-align:center;color:var(--muted);padding:24px;font-style:italic}}
.badge{{display:inline-block;font-size:11px;font-weight:600;padding:2px 9px;border-radius:4px;white-space:nowrap}}
.badge.valid{{background:var(--gbg);color:var(--green)}}.badge.flagged{{background:var(--abg);color:var(--amber)}}
.badge.gap{{background:var(--rbg);color:var(--red)}}
.ft{{margin-top:48px;text-align:center;font-size:12px;color:var(--muted);
border-top:1px solid var(--border);padding-top:20px}}
@media print{{.tw{{overflow:visible}}body{{background:#fff}}}}
</style></head><body><div class="wrap">
<header class="hdr"><h1>Custom Application &mdash; Test Coverage Report</h1>
<p class="sub">Consistency/testability check &amp; three-way traceability (requirement &harr; process step &harr; test case)</p>
<div class="meta">
<div class="mi"><strong>Application</strong>{esc(data.get("application",""))}</div>
<div class="mi"><strong>Process Flow</strong>{esc(flow.get("flow_name","Main flow"))}</div>
<div class="mi"><strong>Run Date</strong>{esc(data.get("run_date",""))}</div>
{baseline_html}
</div></header>

<div class="sc">
<div class="card info"><div class="n">{total_reqs}</div><div class="l">Total Requirements</div></div>
<div class="card good"><div class="n">{testable_reqs}</div><div class="l">Testable</div></div>
<div class="card warn"><div class="n">{flagged_reqs}</div><div class="l">Flagged / Needs Review</div></div>
<div class="card info"><div class="n">{total_flow_steps}</div><div class="l">Process Flow Steps</div></div>
<div class="card good"><div class="n">{total_tcs}</div><div class="l">Test Scenarios Generated</div></div>
<div class="card bad"><div class="n">{len(uncovered_flow_steps)}</div><div class="l">Flow Steps With No Requirement</div></div>
</div>

<h2 class="st">Confirmed Process Flow <span class="pill">REFERENCE</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:90px">Step ID</th><th style="width:18%">Screen / Stage</th>
<th>Description</th><th style="width:22%">Decision Point</th><th style="width:140px">Coverage</th>
</tr></thead><tbody>
{flow_html}
</tbody></table></div>

<h2 class="st">Requirement Validation (Consistency &amp; Testability) <span class="pill">TABLE 1</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:100px">Req ID</th><th style="width:35%">Requirement</th>
<th style="width:110px">Status</th><th style="width:120px">Flow Step(s)</th><th>Notes</th>
</tr></thead><tbody>
{val_html}
</tbody></table></div>

<h2 class="st">Generated Test Scenarios <span class="pill">TABLE 2</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:90px">TC ID</th><th style="width:90px">Req ID</th><th style="width:100px">Flow Step(s)</th>
<th style="width:20%">Title</th><th style="width:16%">Precondition</th>
<th>Steps</th><th style="width:16%">Expected Result</th>
</tr></thead><tbody>
{tc_html}
</tbody></table></div>

<h2 class="st">Three-Way Traceability &mdash; Requirement &harr; Flow Step &harr; Test Case <span class="pill">TABLE 3</span></h2>
<div class="tw"><table><thead><tr>
<th style="width:100px">Req ID</th><th style="width:30%">Requirement</th>
<th style="width:110px">Flow Step(s)</th><th style="width:110px">Linked TC IDs</th><th style="width:160px">Status</th>
</tr></thead><tbody>
{trace_html}
</tbody></table></div>

<div class="ft">Req2QA &mdash; Custom Application Mode &bull; {esc(data.get("run_date",""))}</div>
</div></body></html>"""


def build_xlsx_custom(data, flow, out_path):
    wb = Workbook()
    flow_steps = flow.get("steps", [])
    uncovered_flow_steps = set(data.get("uncovered_flow_steps", []))

    # Sheet 1: Process Flow (reference)
    ws0 = wb.active
    ws0.title = "Process Flow"
    ws0.append(["Step ID", "Screen / Stage", "Description", "Inputs", "Decision Point", "Decision Detail", "Coverage"])
    for st in flow_steps:
        covered = "NO REQUIREMENT" if st.get("step_id") in uncovered_flow_steps else "COVERED"
        ws0.append([
            st.get("step_id", ""), st.get("screen_or_stage", ""), st.get("description", ""),
            ", ".join(st.get("inputs", [])), "YES" if st.get("decision_point") else "NO",
            st.get("decision_detail", ""), covered,
        ])
        r = ws0.max_row
        fill = BAD_FILL if covered == "NO REQUIREMENT" else GOOD_FILL
        for c in range(1, 8):
            ws0.cell(row=r, column=c).fill = fill
            ws0.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws0, 7)
    autosize(ws0, [12, 22, 40, 25, 14, 30, 16])

    # Sheet 2: Validation
    ws1 = wb.create_sheet("Validation")
    ws1.append(["Req ID", "Requirement", "Testable", "Flow Step(s)", "Notes"])
    for v in data.get("validation", []):
        ok = v.get("testable")
        ws1.append([
            v.get("req_id", ""), v.get("requirement", ""), "YES" if ok else "FLAGGED",
            ", ".join(v.get("flow_step_ids", [])), v.get("notes", ""),
        ])
        r = ws1.max_row
        fill = GOOD_FILL if ok else WARN_FILL
        for c in range(1, 6):
            ws1.cell(row=r, column=c).fill = fill
            ws1.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws1, 5)
    autosize(ws1, [12, 50, 14, 18, 40])

    # Sheet 3: Test Scenarios
    ws2 = wb.create_sheet("Test Scenarios")
    ws2.append(["TC ID", "Req ID", "Flow Step(s)", "Title", "Precondition", "Steps", "Expected Result"])
    for s in data.get("test_scenarios", []):
        ws2.append([
            s.get("tc_id", ""), s.get("req_id", ""), ", ".join(s.get("flow_step_ids", [])),
            s.get("title", ""), s.get("precondition", ""),
            s.get("steps", "").replace("\\n", "\n"), s.get("expected_result", ""),
        ])
        r = ws2.max_row
        for c in range(1, 8):
            ws2.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws2, 7)
    autosize(ws2, [10, 10, 16, 30, 26, 45, 30])

    # Sheet 4: Three-Way Traceability Matrix
    ws3 = wb.create_sheet("Traceability Matrix")
    ws3.append(["Req ID", "Requirement", "Flow Step(s)", "Linked TC IDs", "Status"])
    trace = defaultdict(list)
    for s in data.get("test_scenarios", []):
        trace[s.get("req_id", "")].append(s.get("tc_id", ""))
    for v in data.get("validation", []):
        rid = v.get("req_id", "")
        tcs = trace.get(rid, [])
        step_ids = v.get("flow_step_ids", [])
        if not v.get("testable"):
            status, fill = "NOT_TESTABLE", WARN_FILL
        elif not step_ids:
            status, fill = "NO_FLOW_STEP", BAD_FILL
        elif tcs:
            status, fill = "COVERED", GOOD_FILL
        else:
            status, fill = "GAP", BAD_FILL
        ws3.append([rid, v.get("requirement", ""), ", ".join(step_ids), ", ".join(tcs), status])
        r = ws3.max_row
        for c in range(1, 6):
            ws3.cell(row=r, column=c).fill = fill
            ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    step_lookup = {st.get("step_id", ""): st for st in flow_steps}
    for step_id in sorted(uncovered_flow_steps):
        st = step_lookup.get(step_id, {})
        ws3.append(["-", f"{st.get('screen_or_stage', step_id)} - flow step with no requirement", step_id, "-", "UNREQUESTED_STEP"])
        r = ws3.max_row
        for c in range(1, 6):
            ws3.cell(row=r, column=c).fill = BAD_FILL
            ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws3, 5)
    autosize(ws3, [12, 45, 16, 20, 18])

    wb.save(out_path)


def main():
    if len(sys.argv) != 4:
        print("Usage: python3 build_test_scenarios_output.py data.json report_out.html data_out.xlsx")
        sys.exit(1)
    data = json.loads(Path(sys.argv[1]).read_text())
    Path(sys.argv[2]).write_text(build_html(data))
    build_xlsx(data, sys.argv[3])
    print(f"HTML report -> {sys.argv[2]}")
    print(f"Excel data  -> {sys.argv[3]}")


if __name__ == "__main__":
    main()
